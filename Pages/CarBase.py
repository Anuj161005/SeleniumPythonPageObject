import openpyxl
from selenium.webdriver.common.by import By

from Utilities import configReader


class CarBase:
    def __init__(self,driver):
        self.driver = driver

    def getCarTitle(self):
        return self.driver.find_element(By.XPATH,configReader.readConfig("locators","carPageTitle_XPATH")).text


    def carNameAndPrices(self):
        carNames = self.driver.find_elements(By.XPATH,configReader.readConfig("locators","carName_XPATH"))
        carPrices = self.driver.find_elements(By.XPATH,configReader.readConfig("locators","carPrices_XPATH"))

        for i in range (1,len(carPrices)):
            print((carNames[i].text +"-------The Car Prices are-----"+ carPrices[i].text).encode('utf8'))

    # def carNameAndPrices(self,path,sheetname):
    #     workbook = openpyxl.load_workbook(path)
    #     sheet = workbook[sheetname]
    #     carNames = self.driver.find_elements(By.XPATH,configReader.readConfig("locators","carName_XPATH"))
    #     carPrices = self.driver.find_elements(By.XPATH,configReader.readConfig("locators","carPrices_XPATH"))
    #
    #     for i in range (1,len(carPrices)):
    #          sheet.append([carNames[i].text])
    #          sheet.append([carPrices[i].text])
    #     workbook.save(path)

            #print(carNames[i].text +"-------The Car Prices are-----"+ carPrices[i].text)

    # def setCellData(path, sheetname, rowNum, colNum, data):
    #     workbook = openpyxl.load_workbook(path)
    #     sheet = workbook[sheetname]
    #     sheet.cell(row=rowNum, column=colNum, value=data)
    #     workbook.save(path)


