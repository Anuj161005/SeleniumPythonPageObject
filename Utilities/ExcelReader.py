import openpyxl

#from utilities.readingExcel import workbook
#from utilities.writeExcel import sheet


# def getRowCount(path,sheetname):
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook[sheetname]
#     return sheet.max_row
#
# def getColCount(path,sheetname):
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook[sheetname]
#     return sheet.max_column
#
# def getCellData(path,sheetname,rowNum,colNum):
#     workbook = openpyxl.load_workbook(path)
#     sheet = workbook[sheetname]
#     return sheet.cell(row=rowNum,column=colNum).value


def setCellData(path,sheetname,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rowNum,column=colNum,value= data)
    workbook.save(path)

path = "../Excel/testData.xlsx"
sheetname  = "UploadData"

# rows = getRowCount(path,sheetname)
# columns = getColCount(path,sheetname)

# print(rows,"--------",columns)
# print(getCellData(path,sheetname,2,1))

setCellData(path,sheetname,2,1,"DOB")